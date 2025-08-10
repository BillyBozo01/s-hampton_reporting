import os
import sqlite3
import datetime
from uuid import uuid4

from flask import Flask, request, jsonify, send_from_directory, Response, redirect
from openpyxl import Workbook, load_workbook

# ---------------- Paths ----------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FRONTEND_DIR = BASE_DIR  # HTML/CSS/JS live in the project root
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
DB_PATH = os.path.join(BASE_DIR, "reports.db")
EXCEL_PATH = os.path.join(BASE_DIR, "reports.xlsx")

os.makedirs(UPLOAD_DIR, exist_ok=True)

app = Flask(__name__)

# ---------------- Helpers ----------------
def _send_html(name: str):
    path = os.path.join(FRONTEND_DIR, name)
    if not os.path.exists(path):
        return Response("Not found", status=404)
    return send_from_directory(FRONTEND_DIR, name)

def ensure_excel_with_headers():
    """Create the Excel workbook with headers if it doesn't exist."""
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reports"
        ws.append([
            "id", "created_at", "shop_name", "postcode", "details",
            "name", "email", "evidence_paths"
        ])
        wb.save(EXCEL_PATH)

def append_to_excel(row: dict):
    """Append one row (dict) to the workbook."""
    ensure_excel_with_headers()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append([
        row.get("id"),
        row.get("created_at"),
        row.get("shop_name"),
        row.get("postcode"),
        row.get("details"),
        row.get("name"),
        row.get("email"),
        row.get("evidence_paths"),
    ])
    wb.save(EXCEL_PATH)

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# Create table if needed
with get_db() as conn:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            shop_name TEXT,
            postcode TEXT,
            details TEXT NOT NULL,
            name TEXT,
            email TEXT,
            evidence_paths TEXT
        );
        """
    )

# ---------------- API ----------------
@app.post("/api/report")
def api_report():
    # Accept multipart/form-data (FormData) or JSON
    form = request.form if request.form else request.json or {}
    shop_name = (form.get("shop_name") or "").strip()
    postcode = (form.get("postcode") or "").strip()
    details = (form.get("details") or "").strip()
    name = (form.get("name") or "").strip()
    email = (form.get("email") or "").strip()

    if not details:
        return jsonify(ok=False, message="Please include details of the issue."), 400

    saved_files = []
    # Optional file named 'evidence'
    file = request.files.get("evidence")
    if file and file.filename:
        ext = os.path.splitext(file.filename)[1]
        safe_name = f"{uuid4().hex}{ext}"
        dest = os.path.join(UPLOAD_DIR, safe_name)
        file.save(dest)
        saved_files.append(f"/uploads/{safe_name}")

    created_at = datetime.datetime.utcnow().isoformat(timespec="seconds") + "Z"

    with get_db() as conn:
        cur = conn.execute(
            """
            INSERT INTO reports (created_at, shop_name, postcode, details, name, email, evidence_paths)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (created_at, shop_name, postcode, details, name, email, ",".join(saved_files)),
        )
        report_id = cur.lastrowid

    # Mirror to Excel for easy export
    append_to_excel({
        "id": report_id,
        "created_at": created_at,
        "shop_name": shop_name,
        "postcode": postcode,
        "details": details,
        "name": name,
        "email": email,
        "evidence_paths": ",".join(saved_files),
    })

    return jsonify(ok=True, message="Report received. Thank you!")

@app.get("/exports/reports.xlsx")
def download_excel():
    ensure_excel_with_headers()
    return send_from_directory(BASE_DIR, "reports.xlsx", as_attachment=True)

@app.get("/uploads/<path:fname>")
def serve_upload(fname):
    return send_from_directory(UPLOAD_DIR, fname)

# ---------------- Static / Frontend routes ----------------
@app.get("/")
def home():
    # Make reporting.html the homepage
    return redirect("/reporting.html", code=302)

@app.get("/reporting.html")
def reporting_html():
    return _send_html("reporting.html")

# Serve other html pages by name
@app.get("/<page>.html")
def html_pages(page):
    pages = {
        "gdpr": "gdpr.html",
        "tobaccoinfo": "tobaccoinfo.html",
        "vapesinfo": "vapesinfo.html",
        "index": "index.html",  # legacy; will redirect below
    }
    filename = pages.get(page)
    if not filename or not os.path.exists(os.path.join(FRONTEND_DIR, filename)):
        return Response("Not found", status=404)
    if filename == "index.html":
        # Push legacy /index.html to /reporting.html
        return redirect("/reporting.html", code=301)
    return _send_html(filename)

# Legacy clean paths -> .html
@app.get("/reporting")
def reporting_redirect():
    return redirect("/reporting.html", code=301)

@app.get("/gdpr")
def gdpr_redirect():
    return redirect("/gdpr.html", code=301)

@app.get("/tobaccoinfo")
def tob_redirect():
    return redirect("/tobaccoinfo.html", code=301)

@app.get("/vapesinfo")
def vape_redirect():
    return redirect("/vapesinfo.html", code=301)

# Serve CSS/JS assets
@app.get("/style.css")
def style_css():
    return send_from_directory(FRONTEND_DIR, "style.css")

@app.get("/script.js")
def script_js():
    return send_from_directory(FRONTEND_DIR, "script.js")

@app.get("/favicon.ico")
def favicon():
    path = os.path.join(FRONTEND_DIR, "favicon.ico")
    if os.path.exists(path):
        return send_from_directory(FRONTEND_DIR, "favicon.ico")
    return Response(status=204)

# ---------------- Run ----------------
if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
